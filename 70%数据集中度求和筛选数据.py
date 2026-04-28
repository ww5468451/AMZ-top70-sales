import pandas as pd
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Border, Side, Font


def rank100_process(need_process_excel, sheet_name, js_col, date_col, new_gn_excel):
    df = pd.DataFrame(pd.read_excel(need_process_excel, sheet_name=sheet_name,
                                    engine='openpyxl'))
    columns_names = df.columns
    length = len(columns_names)
    # 先将月销售额($)为空值的行去掉，才能不影响后面数据的计算
    df = df.dropna(axis=0, subset=[js_col])
    # print(df)
    df[date_col] = pd.to_datetime(df[date_col]).dt.to_period('M')
    df[js_col] = pd.to_numeric(df[js_col])

    def get_top_100_sales(group):
        return group.sort_values(js_col, ascending=False).head(100)

    # 按月份分组，并将每个分组中销售金额排名前100的商品数据取出来
    top_100_sales = df.groupby([date_col]).apply(get_top_100_sales)
    top_100_sales = top_100_sales.reset_index(drop=True)
    writer = pd.ExcelWriter(new_gn_excel,
                            engine='openpyxl', options={'strings_to_urls': False})
    top_100_sales.to_excel(writer, index=False)
    writer.save()

def percent70_process(need_process_excel, sheet_name, js_col, date_col, new_gn_excel):
    df = pd.DataFrame(pd.read_excel(need_process_excel, sheet_name=sheet_name,
                                    engine='openpyxl'))
    columns_names = df.columns
    length = len(columns_names)
    # 先将月销售额($)为空值的行去掉，才能不影响后面数据的计算
    df = df.dropna(axis=0, subset=[js_col])
    # print(df)
    df[date_col] = pd.to_datetime(df[date_col]).dt.to_period('M')
    df[js_col] = pd.to_numeric(df[js_col])
    monthly_sales = df.groupby(date_col)[js_col].sum()
    # print(grouped)
    # 逐行计算累计销售额并进行筛选
    result = pd.DataFrame()
    def get_top_100_sales(group):
        return group.sort_values(js_col, ascending=False).head(100)
    # 按月份分组，并将每个分组中销售金额排名前100的商品数据取出来
    top_100_sales = df.groupby([date_col]).apply(get_top_100_sales)
    print(top_100_sales.index[0][0])
    for i in range(len(monthly_sales.index)):

        # 筛选每个月的数据
        monthly_top100 = top_100_sales[top_100_sales[date_col] == monthly_sales.index[i]]
        monthly_top100_sum = top_100_sales[top_100_sales[date_col] == monthly_sales.index[i]][js_col].sum()

        monthly_data = df[df[date_col] == monthly_sales.index[i]]
        monthly_data = monthly_data.sort_values(js_col, ascending=False)
        print(monthly_data)
        threshold = monthly_sales[monthly_sales.index[i]] * 0.7
        print(threshold)
        if monthly_top100_sum < threshold:
            cumulative_sales = 0
            for index, row in monthly_data.iterrows():
                result = result.append(row, ignore_index=True)
                cumulative_sales += row[js_col]
                if cumulative_sales > threshold:
                    break
        else:
            monthly_top100_df = pd.DataFrame(monthly_top100)
            result = result.append(monthly_top100_df)
                # 如果加上当前行的月销售额，累计销售额大于阈值，则退出循环
        print('#######################################################')


    result = result.reindex(columns=columns_names)
    writer = pd.ExcelWriter(new_gn_excel,
                            engine='openpyxl', options={'strings_to_urls': False})
    result.to_excel(writer, index=False)
    writer.save()

def css_chage(new_gn_excel):
    # 读取 Excel 文件
    df = pd.read_excel(new_gn_excel, sheet_name='Sheet1')
    columns_names = df.columns
    length = len(columns_names)
    df['日期'] = pd.to_datetime(df['日期'], format='%Y-%m-%d', errors='coerce').dt.date
    writer = pd.ExcelWriter(new_gn_excel,
                            engine='openpyxl', options={'strings_to_urls': False})
    df.to_excel(writer, index=False)
    writer.save()
    writer = pd.ExcelWriter(new_gn_excel, engine='openpyxl',
                            date_format='yyyy/m/d',  options={'strings_to_urls': False})
    try:
        unnamed_idx = df.columns.str.contains('Unnamed')
        df = df.drop(df.columns[unnamed_idx], axis=1)
    except:
        pass

    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    # 打开 Excel 文件并获取工作表对象
    book = openpyxl.load_workbook(new_gn_excel)
    ws = book['Sheet1']
    # 如果只设置了start_color属性而未设置end_color属性，则填充对象将使用start_color属性指定的颜色作为填充的唯一颜色。
    # 如果同时设置了start_color和end_color属性，则填充对象将使用这两个颜色之间的渐变填充单元格。
    orange_fill = PatternFill(start_color='E98A00', end_color='E98A00', fill_type='solid')
    # 设置好单元格的字体颜色，大小，字体类型
    white_font = Font(color='FFFFFF', name='微软雅黑', size=10)
    # 创建边框对象并将其应用于列名单元格，去掉首列的单元格边框
    no_border = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='none'), bottom=Side(style='none'))
    # 还要获取首行包含什么内容，然后获取最大的有内容的单元格，最后进行遍历填充
    for col in range(1, length + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = orange_fill
        cell.font = white_font
        cell.border = no_border

    book.save(new_gn_excel)

if __name__ == '__main__':
    # rank100_process(need_process_excel=r'C:\Users\kyt\Downloads\户外野营桌.xlsx',
    #                   sheet_name='筛选picnic、fold、camp、expand', js_col='月销售额($)', date_col='日期', new_gn_excel='100数据-户外野营桌.xlsx')

    need_process_excel = input('请输入需要筛选文件所在路径（如：C:\\Users\\kyt\\Downloads\\筛选后数据.xlsx）：')
    sheet_name = input('请输入需要筛选文件所在的sheet（如：筛选后数据）：')
    # date_col = input('请输入带有年月的列名（如：时间或者日期）：')
    new_gn_excel = input('请输入筛选后需要保存的路径（如：如：C:\\Users\\kyt\\Downloads\\70%筛选后数据.xlsx）：')
    percent70_process(need_process_excel=need_process_excel,
                      sheet_name=sheet_name, js_col="月销售额($)", date_col="日期", new_gn_excel=new_gn_excel)
    css_chage(new_gn_excel=new_gn_excel)
    # rank100_process(need_process_excel=r'C:\Users\kyt\Downloads\户外野营桌(1).xlsx',
    #                   sheet_name='排除chair、seating等', js_col='月销售额($)', date_col='日D:\0-2024\1月-欧洲站升降电脑桌-市场调研\目标跑70%.xlsx期', new_gn_excel='top100-户外野营桌.xlsx')