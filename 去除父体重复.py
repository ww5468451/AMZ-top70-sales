import pandas as pd
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Border, Side, Font

def PAR_Drop(path, sheet, path1):
    df = pd.read_excel(path, sheet_name=sheet)
    df_max_sales = df.loc[df.groupby(['日期', '父体'])['月销量'].idxmax().dropna()]
    unnamed_idx = df_max_sales.columns.str.contains('Unnamed')
    df_max_sales = df_max_sales.drop(df_max_sales.columns[unnamed_idx], axis=1)
    print(df_max_sales)
    writer = pd.ExcelWriter(path1,
                            engine='openpyxl', engine_kwargs={'options': {'strings_to_urls': False}})
    df_max_sales.to_excel(writer, index=False)
    # 打开 Excel 文件并获取工作表对象
    writer.save()
    book = openpyxl.load_workbook(path1)
    columns_names = df.columns
    length = len(columns_names)
    ws = book['Sheet1']
    # 如果只设置了start_color属性而未设置end_color属性，则填充对象将使用start_color属性指定的颜色作为填充的唯一颜色。
    # 如果同时设置了start_color和end_color属性，则填充对象将使用这两个颜色之间的渐变填充单元格。
    orange_fill = PatternFill(start_color='E98A00', end_color='E98A00', fill_type='solid')
    # 设置好单元格的字体颜色，大小，字体类型
    white_font = Font(color='FFFFFF', name='微软雅黑', size=10)
    # 创建边框对象并将其应用于列名单元格，去掉首列的单元格边框
    no_border = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='none'),
                       bottom=Side(style='none'))
    # 还要获取首行包含什么内容，然后获取最大的有内容的单元格，最后进行遍历填充
    for col in range(1, length + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = orange_fill
        cell.font = white_font
        cell.border = no_border
    book.save(path1)

path = input('请输入需要去除重复父体保留销量最大的文件所在路径（如 C:\\Users\\kyt\\Downloads\\锅具锅盖架数据源(2).xlsx）')
sheet = input('请输入需要去除重复父体保留销量最大的所在的sheet（如：原始数据）')
path1 = input('请输入去除重复父体后的文件所需要保存的路径（如：C:\\Users\\kyt\\Downloads\\锅具锅盖架数据源去重父体.xlsx）')
PAR_Drop(path=path, sheet=sheet, path1=path1)
